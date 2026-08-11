#!/usr/bin/env python3
"""Import temperature readings from CSV or Excel into NOVARAReadings.

Expected columns (header names are case-insensitive; aliases accepted):

  Required:
    TimestampUTC  — ISO-8601 UTC preferred (aliases: Timestamp, Timestamp (UTC), DateTime, Time)
    T1            — supply temperature °F (aliases: t1, T1 (F), Supply, SupplyTemp)
    T2            — return temperature °F (aliases: t2, T2 (F), Return, ReturnTemp)

  Site / system (column, CLI flag, or filename for Vista Springs):
    SiteID        — e.g. SITE001 (aliases: Site Id, site_id, Site)
                    or pass --site-id / --default-site-id
    SystemID      — e.g. SYS001 (aliases: System Id, system_id, System)
                    or pass --system-id / --default-system-id

  Optional:
    RelayState    — numeric relay state (aliases: Relay, Relay State, relay_state)

Place files under data/readings/ (recommended), then run:

  python3 scripts/import_readings.py data/readings/my_export.csv --dry-run \\
    --site-id SITE001 --system-id SYS001
  python3 scripts/import_readings.py data/readings/my_export.csv --execute \\
    --site-id SITE001 --system-id SYS001

Import all Vista Springs DHW exports in one run (DHW-Sys-N → SITE001 / SYS00N):

  python3 scripts/import_readings.py data/readings/vista-springs --dry-run
  python3 scripts/import_readings.py data/readings/vista-springs --execute

CSV may omit SiteID/SystemID when those flags (or Vista Springs filenames) are provided:

  TimestampUTC,T1,T2,RelayState
  Timestamp (UTC),Relay State,T2 (F),T1 (F)

DynamoDB keys are SiteID (HASH) + TimestampUTC (RANGE). When SystemID is set,
TimestampUTC is stored as ``{iso}#{SystemID}`` so multiple systems at the same
site/time can coexist; the API strips the suffix for charts. SystemID is also
stored on each item for filtering. By default, existing keys are skipped so
re-imports are safe. Use --overwrite to replace values.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from collections import defaultdict
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

# Allow importing novara_api helpers when run from repo root or scripts/.
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)

from novara_api import (  # noqa: E402
    TABLE_NAME,
    dynamodb_resource,
    ensure_readings_table,
    reading_sort_key,
    sanitize_aws_env,
)

SITE_ID_RE = re.compile(r"^SITE\d{3,}$", re.IGNORECASE)
SYSTEM_ID_RE = re.compile(r"^SYS\d{3,}$", re.IGNORECASE)
# Vista Springs exports: DHW-Sys-1-chart-data-....csv.csv → SITE001 / SYS001
DHW_SYS_FILE_RE = re.compile(r"DHW-Sys-(\d+)", re.IGNORECASE)
READING_FILE_SUFFIXES = (".csv", ".txt", ".tsv", ".xlsx", ".xlsm")
VISTA_SPRINGS_SITE_ID = "SITE001"

COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "SiteID": (
        "siteid",
        "site_id",
        "site id",
        "site",
        "siteidentifier",
        "site_identifier",
    ),
    "SystemID": (
        "systemid",
        "system_id",
        "system id",
        "system",
        "systemidentifier",
        "system_identifier",
    ),
    "TimestampUTC": (
        "timestamputc",
        "timestamp_utc",
        "timestamp utc",
        "timestamp",
        "datetime",
        "date_time",
        "date time",
        "time",
        "date",
        "readingtime",
        "reading_time",
        "recordedat",
        "recorded_at",
    ),
    "T1": ("t1", "supply", "supplytemp", "supply_temp", "supply temperature", "temp1"),
    "T2": ("t2", "return", "returntemp", "return_temp", "return temperature", "temp2"),
    "RelayState": (
        "relaystate",
        "relay_state",
        "relay state",
        "relay",
        "relaystatus",
        "relay_status",
    ),
}


def normalize_header(name: str) -> str:
    """Normalize CSV headers; strip units in parentheses (e.g. T1 (F), Timestamp (UTC))."""
    text = (name or "").strip().lower()
    text = re.sub(r"\([^)]*\)", " ", text)
    text = text.replace("-", " ").replace("/", " ").replace("_", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def is_reading_file(path: Path) -> bool:
    name = path.name.lower()
    if name.startswith("."):
        return False
    return any(name.endswith(suffix) for suffix in READING_FILE_SUFFIXES)


def expand_input_paths(paths: list[Path]) -> list[Path]:
    """Expand directories to reading files (sorted); keep explicit files as-is."""
    expanded: list[Path] = []
    for path in paths:
        if path.is_dir():
            files = sorted(p for p in path.iterdir() if p.is_file() and is_reading_file(p))
            if not files:
                raise ValueError(f"No reading files found in directory: {path}")
            expanded.extend(files)
        else:
            expanded.append(path)
    return expanded


def infer_vista_springs_ids(path: Path) -> tuple[str, str] | None:
    """Map DHW-Sys-N-... filenames to (SITE001, SYS00N)."""
    match = DHW_SYS_FILE_RE.search(path.name)
    if not match:
        return None
    system_num = int(match.group(1))
    return VISTA_SPRINGS_SITE_ID, f"SYS{system_num:03d}"


def resolve_columns(headers: list[str]) -> dict[str, str]:
    """Map canonical field -> actual header name present in the file."""
    by_norm = {normalize_header(h): h for h in headers if str(h).strip()}
    resolved: dict[str, str] = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in by_norm:
                resolved[canonical] = by_norm[alias]
                break
    return resolved


def parse_site_id(value: Any, site_map: dict[str, str], default_site: str | None) -> str:
    raw = "" if value is None else str(value).strip()
    if not raw:
        if default_site:
            return default_site.upper()
        raise ValueError("SiteID is empty (provide column or --site-id)")
    mapped = site_map.get(raw) or site_map.get(raw.upper())
    site_id = (mapped or raw).strip().upper()
    if not SITE_ID_RE.match(site_id):
        raise ValueError(
            f"SiteID '{raw}' is not SITE### "
            f"(map it with --site-map {raw}=SITE001 or use that form in the file)"
        )
    return site_id


def parse_system_id(value: Any, default_system: str | None) -> str | None:
    """Return normalized SystemID, optional default, or None when omitted."""
    raw = "" if value is None else str(value).strip()
    if not raw:
        if default_system:
            return default_system.upper()
        return None
    system_id = raw.upper()
    if not SYSTEM_ID_RE.match(system_id):
        raise ValueError(
            f"SystemID '{raw}' is not SYS### "
            "(use SYS001 in the file or pass --system-id SYS001)"
        )
    return system_id


def _excel_serial_to_utc(serial: float) -> datetime:
    # Excel serial date: days since 1899-12-30 (Windows / openpyxl convention).
    from datetime import timedelta

    epoch = datetime(1899, 12, 30, tzinfo=timezone.utc)
    return epoch + timedelta(days=float(serial))


def parse_timestamp_utc(value: Any) -> str:
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError("TimestampUTC is empty")

    if isinstance(value, datetime):
        dt = value
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        return dt.strftime("%Y-%m-%dT%H:%M:%SZ")

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Likely an Excel serial date.
        if 20000 < float(value) < 80000:
            return _excel_serial_to_utc(float(value)).strftime("%Y-%m-%dT%H:%M:%SZ")
        raise ValueError(f"Unrecognized numeric timestamp: {value}")

    text = str(value).strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"

    # Common export forms: "2026-08-01 14:30:00" / "2026-08-01T14:30:00"
    candidates = [text]
    if " " in text and "T" not in text:
        candidates.append(text.replace(" ", "T", 1))

    for candidate in candidates:
        try:
            dt = datetime.fromisoformat(candidate)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            else:
                dt = dt.astimezone(timezone.utc)
            return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            pass

    for fmt in (
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m/%d/%Y",
        "%Y-%m-%d",
    ):
        try:
            dt = datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
            return dt.strftime("%Y-%m-%dT%H:%M:%SZ")
        except ValueError:
            continue

    raise ValueError(f"Unrecognized timestamp: {value!r}")


def parse_number(value: Any, field: str, *, required: bool = True) -> Decimal | None:
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            raise ValueError(f"{field} is empty")
        return None
    if isinstance(value, bool):
        raise ValueError(f"{field} must be a number")
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"{field} must be a number (got {value!r})") from exc


def parse_site_map(pairs: list[str] | None) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for pair in pairs or []:
        if "=" not in pair:
            raise ValueError(f"Invalid --site-map entry '{pair}' (use OLD=SITE001)")
        old, new = pair.split("=", 1)
        old = old.strip()
        new = new.strip().upper()
        if not old or not SITE_ID_RE.match(new):
            raise ValueError(f"Invalid --site-map entry '{pair}' (target must be SITE###)")
        mapping[old] = new
        mapping[old.upper()] = new
    return mapping


def row_to_item(
    row: dict[str, Any],
    columns: dict[str, str],
    *,
    site_map: dict[str, str],
    default_site: str | None,
    default_system: str | None,
) -> dict[str, Any]:
    missing = [name for name in ("TimestampUTC", "T1", "T2") if name not in columns]
    if "SiteID" not in columns and not default_site:
        missing.append("SiteID")
    if missing:
        raise ValueError(f"Missing required column(s): {', '.join(missing)}")

    site_raw = row.get(columns["SiteID"]) if "SiteID" in columns else None
    system_raw = row.get(columns["SystemID"]) if "SystemID" in columns else None
    item: dict[str, Any] = {
        "SiteID": parse_site_id(site_raw, site_map, default_site),
        "TimestampUTC": parse_timestamp_utc(row.get(columns["TimestampUTC"])),
        "T1": parse_number(row.get(columns["T1"]), "T1"),
        "T2": parse_number(row.get(columns["T2"]), "T2"),
    }
    system_id = parse_system_id(system_raw, default_system)
    if system_id:
        item["SystemID"] = system_id
        # Composite sort key so SYS001 and SYS002 at the same site/time do not collide.
        item["TimestampUTC"] = reading_sort_key(item["TimestampUTC"], system_id)
    if "RelayState" in columns:
        relay = parse_number(row.get(columns["RelayState"]), "RelayState", required=False)
        if relay is not None:
            item["RelayState"] = relay
    return item


def _read_csv_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        if not reader.fieldnames:
            raise ValueError("CSV has no header row")
        headers = [str(h) for h in reader.fieldnames]
        rows = [dict(row) for row in reader]
    return headers, rows


def _read_excel_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Excel support requires openpyxl. Install with: pip install openpyxl\n"
            "Or export the sheet as CSV and import that instead."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("Excel sheet is empty") from exc

    headers = ["" if h is None else str(h).strip() for h in header_row]
    if not any(headers):
        raise ValueError("Excel sheet has no header row")

    rows: list[dict[str, Any]] = []
    for values in rows_iter:
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        row: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row[header] = values[idx] if idx < len(values) else None
        rows.append(row)
    return headers, rows


def load_rows(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt", ".tsv"}:
        return _read_csv_rows(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_excel_rows(path)
    raise ValueError(
        f"Unsupported file type '{suffix}'. Use .csv or .xlsx "
        "(legacy .xls is not supported — re-save as .xlsx or .csv)."
    )


def _cell_blank(row: dict[str, Any], header: str | None) -> bool:
    if not header:
        return True
    value = row.get(header)
    return value is None or (isinstance(value, str) and not value.strip())


def parse_items(
    path: Path,
    *,
    site_map: dict[str, str],
    default_site: str | None,
    default_system: str | None = None,
) -> tuple[list[dict[str, Any]], list[str]]:
    headers, rows = load_rows(path)
    columns = resolve_columns(headers)
    items: list[dict[str, Any]] = []
    errors: list[str] = []
    for index, row in enumerate(rows, start=2):  # header is row 1
        if not any(str(v).strip() for v in row.values() if v is not None):
            continue
        # Chart exports often pad future timestamps with blank T1/T2 — skip those.
        t1_blank = _cell_blank(row, columns.get("T1"))
        t2_blank = _cell_blank(row, columns.get("T2"))
        if t1_blank and t2_blank:
            continue
        try:
            items.append(
                row_to_item(
                    row,
                    columns,
                    site_map=site_map,
                    default_site=default_site,
                    default_system=default_system,
                )
            )
        except ValueError as exc:
            errors.append(f"row {index}: {exc}")
    return items, errors


def dedupe_items(items: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    """Keep the last occurrence of each SiteID(+SystemID)+TimestampUTC within the file."""
    by_key: dict[tuple[str, str, str], dict[str, Any]] = {}
    for item in items:
        key = (
            item["SiteID"],
            str(item.get("SystemID") or ""),
            item["TimestampUTC"],
        )
        by_key[key] = item
    return list(by_key.values()), len(items) - len(by_key)


def put_items(
    table,
    items: Iterable[dict[str, Any]],
    *,
    overwrite: bool,
    dry_run: bool,
) -> dict[str, int]:
    written = 0
    skipped = 0
    overwritten = 0
    errors = 0

    for item in items:
        if dry_run:
            written += 1
            continue
        try:
            if overwrite:
                table.put_item(Item=item)
                overwritten += 1
            else:
                table.put_item(
                    Item=item,
                    ConditionExpression=(
                        "attribute_not_exists(SiteID) AND attribute_not_exists(TimestampUTC)"
                    ),
                )
                written += 1
        except Exception as exc:  # noqa: BLE001 — treat conditional failures as skips
            error_code = ""
            response = getattr(exc, "response", None)
            if isinstance(response, dict):
                error_code = str((response.get("Error") or {}).get("Code") or "")
            if (
                error_code == "ConditionalCheckFailedException"
                or type(exc).__name__ == "ConditionalCheckFailedException"
                or "ConditionalCheckFailed" in str(exc)
            ):
                skipped += 1
                continue
            errors += 1
            print(
                f"ERROR writing {item.get('SiteID')}/{item.get('TimestampUTC')}: {exc}",
                file=sys.stderr,
            )

    return {
        "written": written,
        "skipped": skipped,
        "overwritten": overwritten,
        "errors": errors,
    }


def summarize_by_system(items: Iterable[dict[str, Any]]) -> dict[str, int]:
    counts: dict[str, int] = defaultdict(int)
    for item in items:
        counts[str(item.get("SystemID") or "(none)")] += 1
    return dict(sorted(counts.items()))


def print_system_summary(
    label: str,
    items: list[dict[str, Any]],
    *,
    stats_by_system: dict[str, dict[str, int]] | None = None,
) -> None:
    counts = summarize_by_system(items)
    print(f"{label}:")
    if not counts:
        print("  (no records)")
        return
    for system_id, count in counts.items():
        if stats_by_system and system_id in stats_by_system:
            stats = stats_by_system[system_id]
            if "overwritten" in stats and stats.get("overwritten", 0) and not stats.get(
                "written", 0
            ):
                detail = (
                    f"overwrote {stats['overwritten']}, errors={stats['errors']}"
                )
            else:
                detail = (
                    f"wrote {stats['written']}, skipped {stats['skipped']}, "
                    f"errors={stats['errors']}"
                )
            print(f"  {system_id}: {count} record(s) → {detail}")
        else:
            print(f"  {system_id}: {count} record(s)")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "files",
        nargs="+",
        type=Path,
        help=(
            "CSV/Excel files or a directory of them "
            "(e.g. data/readings/vista-springs)"
        ),
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate without writing to DynamoDB",
    )
    parser.add_argument(
        "--execute",
        action="store_true",
        help="Write items to DynamoDB (required unless --dry-run)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace existing SiteID+TimestampUTC items (default: skip duplicates)",
    )
    parser.add_argument(
        "--site-id",
        "--default-site-id",
        dest="default_site_id",
        help="SiteID when the file has no SiteID column (e.g. SITE001)",
    )
    parser.add_argument(
        "--system-id",
        "--default-system-id",
        dest="default_system_id",
        help="SystemID to store on each reading when the file has no SystemID column (e.g. SYS001)",
    )
    parser.add_argument(
        "--site-map",
        action="append",
        default=[],
        metavar="OLD=SITE001",
        help="Map a source site identifier to SITE### (repeatable)",
    )
    parser.add_argument(
        "--table",
        default=os.environ.get("NOVARA_READINGS_TABLE", TABLE_NAME),
        help=f"DynamoDB table name (default: {TABLE_NAME})",
    )
    args = parser.parse_args(argv)

    if not args.dry_run and not args.execute:
        parser.error("Pass --dry-run or --execute")

    try:
        site_map = parse_site_map(args.site_map)
    except ValueError as exc:
        parser.error(str(exc))

    default_site = args.default_site_id.strip().upper() if args.default_site_id else None
    if default_site and not SITE_ID_RE.match(default_site):
        parser.error("--site-id / --default-site-id must look like SITE001")

    default_system = (
        args.default_system_id.strip().upper() if args.default_system_id else None
    )
    if default_system and not SYSTEM_ID_RE.match(default_system):
        parser.error("--system-id / --default-system-id must look like SYS001")

    try:
        input_paths = expand_input_paths(list(args.files))
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    all_items: list[dict[str, Any]] = []
    had_errors = False
    for path in input_paths:
        if not path.is_file():
            print(f"ERROR: file not found: {path}", file=sys.stderr)
            had_errors = True
            continue

        file_site = default_site
        file_system = default_system
        inferred = infer_vista_springs_ids(path)
        if inferred:
            inferred_site, inferred_system = inferred
            if not file_site:
                file_site = inferred_site
            if not file_system:
                file_system = inferred_system

        print(f"Reading {path} …")
        if inferred and (file_site or file_system):
            print(f"  mapped → SiteID={file_site} SystemID={file_system}")

        try:
            items, errors = parse_items(
                path,
                site_map=site_map,
                default_site=file_site,
                default_system=file_system,
            )
        except (ValueError, RuntimeError) as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            had_errors = True
            continue
        for err in errors[:20]:
            print(f"  {err}", file=sys.stderr)
        if len(errors) > 20:
            print(f"  … and {len(errors) - 20} more row errors", file=sys.stderr)
        if errors:
            had_errors = True
        print(f"  parsed {len(items)} row(s), {len(errors)} error(s)")
        all_items.extend(items)

    if not all_items:
        print("No valid readings to import.")
        return 1 if had_errors else 0

    unique_items, file_dupes = dedupe_items(all_items)
    if file_dupes:
        print(f"Collapsed {file_dupes} duplicate key(s) within input (kept last).")

    sites = sorted({item["SiteID"] for item in unique_items})
    systems = sorted({item["SystemID"] for item in unique_items if item.get("SystemID")})
    systems_label = ", ".join(systems) if systems else "(none)"
    print(
        f"Ready: {len(unique_items)} item(s) across site(s) {', '.join(sites)} "
        f"system(s) {systems_label} → table {args.table}"
    )
    print_system_summary("Records per system (parsed)", unique_items)
    if unique_items:
        sample = unique_items[0]
        sample_out = {
            "SiteID": sample["SiteID"],
            "TimestampUTC": sample["TimestampUTC"],
            "T1": float(sample["T1"]),
            "T2": float(sample["T2"]),
        }
        if sample.get("SystemID"):
            sample_out["SystemID"] = sample["SystemID"]
        if "RelayState" in sample:
            sample_out["RelayState"] = float(sample["RelayState"])
        print("Sample item:", sample_out)

    if args.dry_run:
        print("[DRY-RUN] No writes performed.")
        return 1 if had_errors else 0

    sanitize_aws_env()
    if args.table == TABLE_NAME:
        ensure_readings_table()
    table = dynamodb_resource().Table(args.table)

    # Write per system so the summary matches DynamoDB outcomes clearly.
    by_system: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for item in unique_items:
        by_system[str(item.get("SystemID") or "(none)")].append(item)

    stats_by_system: dict[str, dict[str, int]] = {}
    totals = {"written": 0, "skipped": 0, "overwritten": 0, "errors": 0}
    for system_id in sorted(by_system):
        stats = put_items(
            table,
            by_system[system_id],
            overwrite=args.overwrite,
            dry_run=False,
        )
        stats_by_system[system_id] = stats
        for key in totals:
            totals[key] += stats[key]

    print_system_summary(
        "Import summary per system",
        unique_items,
        stats_by_system=stats_by_system,
    )
    if args.overwrite:
        print(
            f"Total: wrote/overwrote {totals['overwritten']} item(s); "
            f"errors={totals['errors']}"
        )
    else:
        print(
            f"Total: wrote {totals['written']} new item(s); "
            f"skipped {totals['skipped']} existing; errors={totals['errors']}"
        )
    if totals["errors"] or had_errors:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
